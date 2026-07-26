from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None
    Border = Font = PatternFill = Side = None
    get_column_letter = None
    EXCEL_AVAILABLE = False


EXPERIMENT_NAME = "扰动列车数灵敏度实验"
BASE_SCRIPT_NAME = "run_frankfurt_hbf_benchmark.py"
DEFAULT_OUTPUT_DIR = "frankfurt_hbf_source_count_sensitivity"
SUMMARY_FILE = "法兰克福中央车站QEA-NS与CP-SAT对比实验汇总.csv"
DETAIL_FILE = "法兰克福中央车站QEA-NS与CP-SAT方案指标对比.csv"
PERFORMANCE_FILE = "法兰克福中央车站QEA-NS与CP-SAT求解性能对比.csv"
REQUIRED_OUTPUT_FILES = (SUMMARY_FILE, DETAIL_FILE, PERFORMANCE_FILE)
METHOD_ORDER = ["QEA-NS", "CP-SAT"]
DISTURBANCE_EVENT_COUNT = 2
DETAIL_FIELDS = [
    "实验名称",
    "seed",
    "每事件目标源头列车数",
    "场景实际源头列车总数",
    "源头列车",
    "源头列车评分",
    "扰动继承晚点列车数",
    "扰动继承晚点合计",
    "方法",
    "状态",
    "总冲突项",
    "硬冲突项",
    "进站咽喉冲突项",
    "股道占用冲突项",
    "出站咽喉冲突项",
    "总晚点时长",
    "平均晚点时长",
    "最大晚点时长",
    "能量值",
    "求解时间(s)",
    "内生顺延列车数",
    "内生顺延合计(min)",
]
COMPARISON_FIELDS = [
    "实验名称",
    "seed",
    "每事件目标源头列车数",
    "场景实际源头列车总数",
    "源头列车",
    "扰动继承晚点列车数",
    "扰动继承晚点合计",
    "QEA-NS硬冲突",
    "CP-SAT硬冲突",
    "QEA-NS总晚点",
    "CP-SAT总晚点",
    "QEA-NS相对CP-SAT晚点优势",
    "QEA-NS求解时间(s)",
    "CP-SAT求解时间(s)",
    "结论",
]


def parse_count_spec(count_spec: str) -> list[int]:
    counts: list[int] = []
    for part in count_spec.split(","):
        token = part.strip()
        if not token:
            continue
        value = int(token)
        if value < 1:
            raise ValueError("源头列车数必须为正整数。")
        counts.append(value)
    return list(dict.fromkeys(counts))


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def to_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "--":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_summary(case_dir: Path) -> dict[str, str]:
    rows = read_csv_rows(case_dir / SUMMARY_FILE)
    return {row.get("项目", ""): row.get("值", "") for row in rows}


def disturbance_event_count(summary: dict[str, str]) -> int:
    text = str(summary.get("扰动时刻", "")).strip()
    if not text:
        return 0
    return len([part for part in re.split(r"[;,；，、\s]+", text) if part])


def scenario_source_train_total(summary: dict[str, str]) -> int:
    return int(float(summary.get("源头列车数", "")))


def expected_source_train_total(
    summary: dict[str, str], per_event_count: int
) -> int | None:
    if disturbance_event_count(summary) != DISTURBANCE_EVENT_COUNT:
        return None
    return per_event_count * DISTURBANCE_EVENT_COUNT


def actual_source_train_count_per_event(summary: dict[str, str]) -> int:
    event_count = disturbance_event_count(summary)
    actual_total = scenario_source_train_total(summary)
    if event_count != DISTURBANCE_EVENT_COUNT:
        raise RuntimeError(
            f"扰动事件数不一致：要求{DISTURBANCE_EVENT_COUNT}个，实际{event_count}个"
        )
    if actual_total % event_count:
        raise RuntimeError(
            f"场景源头列车总数{actual_total}不能按{event_count}个事件平均解释"
        )
    return actual_total // event_count


def read_detail_metrics(case_dir: Path) -> dict[str, dict[str, str]]:
    detail_rows = read_csv_rows(case_dir / DETAIL_FILE)
    metrics_by_method = {method: {} for method in METHOD_ORDER}
    for row in detail_rows:
        metric = row.get("指标", "")
        for method in METHOD_ORDER:
            metrics_by_method[method][metric] = row.get(method, "--")
    return metrics_by_method


def read_performance_metrics(case_dir: Path) -> dict[str, dict[str, str]]:
    rows = read_csv_rows(case_dir / PERFORMANCE_FILE)
    return {row.get("方法", ""): row for row in rows}


def case_completed(case_dir: Path, source_count: int) -> bool:
    if not all((case_dir / filename).exists() for filename in REQUIRED_OUTPUT_FILES):
        return False
    summary = read_summary(case_dir)
    performance_methods = {
        row.get("方法", "") for row in read_csv_rows(case_dir / PERFORMANCE_FILE)
    }
    try:
        actual_total = scenario_source_train_total(summary)
    except (TypeError, ValueError):
        return False
    expected_total = expected_source_train_total(summary, source_count)
    return (
        expected_total is not None
        and actual_total == expected_total
        and set(METHOD_ORDER).issubset(performance_methods)
    )


def cleanup_case_outputs(case_dir: Path) -> None:
    if not case_dir.exists():
        return
    for path in case_dir.iterdir():
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()


def build_command(
    args: argparse.Namespace, source_count: int, case_dir: Path
) -> list[str]:
    command = [
        sys.executable,
        str(Path(__file__).resolve().parent / BASE_SCRIPT_NAME),
        "--seed",
        str(args.seed),
        "--source-train-count",
        str(source_count),
        "--output-dir",
        str(case_dir),
        "--max-route-candidates",
        str(args.max_route_candidates),
        "--qea-time-limit",
        str(args.qea_time_limit),
        "--subproblem-time-limit",
        str(args.subproblem_time_limit),
        "--cp-sat-time-limit",
        str(args.cp_sat_time_limit),
        "--qea-pop-size",
        str(args.qea_pop_size),
        "--qea-max-generations",
        str(args.qea_max_generations),
        "--qea-neighborhood-restarts",
        str(args.qea_neighborhood_restarts),
        "--safety-wait-rounds",
        str(args.safety_wait_rounds),
    ]
    if args.short_case_train_limit:
        command.extend(["--short-case-train-limit", str(args.short_case_train_limit)])
    if args.input:
        command.extend(["--input", args.input])
    return command


def run_case_command(command: list[str]) -> int:
    env = os.environ.copy()
    env["PYTHONHASHSEED"] = "0"
    process = subprocess.Popen(
        command,
        cwd=Path(__file__).resolve().parent,
        stderr=subprocess.STDOUT,
        text=True,
        env=env,
    )
    try:
        return process.wait()
    except KeyboardInterrupt:
        print("用户中断，正在终止当前灵敏度实验。")
        process.terminate()
        try:
            process.wait(timeout=30)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait()
        return 130


def validate_source_count(case_dir: Path, source_count: int) -> None:
    summary = read_summary(case_dir)
    try:
        actual_total = scenario_source_train_total(summary)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"无法读取场景源头列车总数：{case_dir}") from exc
    expected_total = expected_source_train_total(summary, source_count)
    if expected_total is None:
        raise RuntimeError(
            f"扰动事件数不一致：要求{DISTURBANCE_EVENT_COUNT}个，"
            f"实际{disturbance_event_count(summary)}个，目录={case_dir}"
        )
    if actual_total != expected_total:
        raise RuntimeError(
            f"场景源头列车总数不一致：每事件目标{source_count}，"
            f"事件数{DISTURBANCE_EVENT_COUNT}，应为{expected_total}，"
            f"实际{actual_total}，目录={case_dir}"
        )


def collect_case_result(
    seed: int, source_count: int, case_dir: Path
) -> list[dict[str, object]]:
    summary = read_summary(case_dir)
    detail = read_detail_metrics(case_dir)
    performance = read_performance_metrics(case_dir)
    actual_total = scenario_source_train_total(summary)
    rows: list[dict[str, object]] = []
    for method in METHOD_ORDER:
        method_detail = detail.get(method, {})
        method_perf = performance.get(method, {})
        rows.append(
            {
                "实验名称": EXPERIMENT_NAME,
                "seed": seed,
                "每事件目标源头列车数": source_count,
                "场景实际源头列车总数": actual_total,
                "源头列车": summary.get("源头列车", "--"),
                "源头列车评分": summary.get("源头列车评分", "--"),
                "扰动继承晚点列车数": summary.get("扰动继承晚点列车数", "--"),
                "扰动继承晚点合计": summary.get("扰动继承晚点合计", "--"),
                "方法": method,
                "状态": method_perf.get("状态", "--"),
                "总冲突项": method_detail.get("总冲突项", "--"),
                "硬冲突项": method_detail.get("硬冲突项", "--"),
                "进站咽喉冲突项": method_detail.get("进站咽喉冲突项", "--"),
                "股道占用冲突项": method_detail.get("股道占用冲突项", "--"),
                "出站咽喉冲突项": method_detail.get("出站咽喉冲突项", "--"),
                "总晚点时长": method_detail.get("总晚点时长", "--"),
                "平均晚点时长": method_detail.get("平均晚点时长", "--"),
                "最大晚点时长": method_detail.get("最大晚点时长", "--"),
                "能量值": method_detail.get("能量值", "--"),
                "求解时间(s)": method_perf.get("求解时间(s)", "--"),
                "内生顺延列车数": method_perf.get("内生顺延列车数", "--"),
                "内生顺延合计(min)": method_perf.get("内生顺延合计(min)", "--"),
            }
        )
    return rows


def find_row(
    rows: list[dict[str, object]], source_count: int, method: str
) -> dict[str, object] | None:
    for row in rows:
        if row.get("每事件目标源头列车数") == source_count and row.get("方法") == method:
            return row
    return None


def build_comparison_rows(
    rows: list[dict[str, object]], source_counts: list[int]
) -> list[dict[str, object]]:
    comparison_rows: list[dict[str, object]] = []
    for source_count in source_counts:
        qea_row = find_row(rows, source_count, "QEA-NS") or {}
        cp_sat_row = find_row(rows, source_count, "CP-SAT") or {}
        qea_delay = to_number(qea_row.get("总晚点时长"))
        cp_sat_delay = to_number(cp_sat_row.get("总晚点时长"))
        advantage = (
            cp_sat_delay - qea_delay
            if qea_delay is not None and cp_sat_delay is not None
            else None
        )
        if advantage is None:
            conclusion = "不可比较"
        elif advantage > 0:
            conclusion = "QEA-NS总晚点更低"
        elif advantage < 0:
            conclusion = "CP-SAT总晚点更低"
        else:
            conclusion = "总晚点持平"
        comparison_rows.append(
            {
                "实验名称": EXPERIMENT_NAME,
                "seed": qea_row.get("seed", cp_sat_row.get("seed", "--")),
                "每事件目标源头列车数": source_count,
                "场景实际源头列车总数": qea_row.get("场景实际源头列车总数", "--"),
                "源头列车": qea_row.get("源头列车", "--"),
                "扰动继承晚点列车数": qea_row.get("扰动继承晚点列车数", "--"),
                "扰动继承晚点合计": qea_row.get("扰动继承晚点合计", "--"),
                "QEA-NS硬冲突": qea_row.get("硬冲突项", "--"),
                "CP-SAT硬冲突": cp_sat_row.get("硬冲突项", "--"),
                "QEA-NS总晚点": qea_row.get("总晚点时长", "--"),
                "CP-SAT总晚点": cp_sat_row.get("总晚点时长", "--"),
                "QEA-NS相对CP-SAT晚点优势": round(advantage, 3)
                if advantage is not None
                else "--",
                "QEA-NS求解时间(s)": qea_row.get("求解时间(s)", "--"),
                "CP-SAT求解时间(s)": cp_sat_row.get("求解时间(s)", "--"),
                "结论": conclusion,
            }
        )
    return comparison_rows


def write_excel(path: Path, sheets: list[tuple[str, list[dict[str, object]]]]) -> None:
    if not EXCEL_AVAILABLE:
        return
    assert openpyxl is not None
    assert Border is not None
    assert Font is not None
    assert PatternFill is not None
    assert Side is not None
    assert get_column_letter is not None
    workbook = openpyxl.Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for index, (sheet_name, rows) in enumerate(sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        if worksheet is None:
            worksheet = workbook.create_sheet(sheet_name)
        worksheet.title = sheet_name
        if not rows:
            continue
        fieldnames = list(rows[0])
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([row.get(field, "") for field in fieldnames])
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                cell.border = border
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_index = int(column_cells[0].column or 1)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12), 42
            )
    workbook.save(path)


def write_outputs(
    root_dir: Path, rows: list[dict[str, object]], source_counts: list[int]
) -> None:
    comparison_rows = build_comparison_rows(rows, source_counts)
    write_csv(root_dir / "扰动列车数灵敏度逐次结果.csv", rows, DETAIL_FIELDS)
    write_csv(
        root_dir / "扰动列车数灵敏度对比汇总.csv", comparison_rows, COMPARISON_FIELDS
    )
    write_excel(
        root_dir / "扰动列车数灵敏度实验汇总.xlsx",
        [
            ("逐次结果", rows),
            ("对比汇总", comparison_rows),
        ],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=EXPERIMENT_NAME)
    parser.add_argument("--source-counts", type=str, default="3,6,9,12,15")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--input", type=str, default="")
    parser.add_argument("--output-dir", type=str, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--keep-existing", action="store_true")
    parser.add_argument("--rerun-completed", action="store_true")
    parser.add_argument("--max-route-candidates", type=int, default=5)
    parser.add_argument("--qea-time-limit", type=float, default=400.0)
    parser.add_argument("--subproblem-time-limit", type=float, default=120.0)
    parser.add_argument(
        "--cp-sat-time-limit", dest="cp_sat_time_limit", type=float, default=400.0
    )
    parser.add_argument("--qea-pop-size", type=int, default=50)
    parser.add_argument("--qea-max-generations", type=int, default=500)
    parser.add_argument("--qea-neighborhood-restarts", type=int, default=1)
    parser.add_argument("--safety-wait-rounds", type=int, default=160)
    parser.add_argument("--short-case-train-limit", type=int, default=0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_counts = parse_count_spec(args.source_counts)
    if not source_counts:
        raise ValueError("至少需要一个源头列车数。")

    root_dir = Path(args.output_dir).expanduser().resolve()
    root_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, object]] = []

    for source_count in source_counts:
        case_dir = root_dir / f"source_count_{source_count}"
        case_dir.mkdir(parents=True, exist_ok=True)
        return_code = 0
        if case_completed(case_dir, source_count) and not args.rerun_completed:
            print(f"源头列车数={source_count} 已完成，跳过。")
        else:
            if not args.keep_existing:
                cleanup_case_outputs(case_dir)
            command = build_command(args, source_count, case_dir)
            print(f"开始源头列车数={source_count}")
            return_code = run_case_command(command)
            if return_code != 0:
                case_rows = collect_case_result(args.seed, source_count, case_dir)
                all_rows.extend(case_rows)
                write_outputs(root_dir, all_rows, source_counts)
                raise RuntimeError(
                    f"源头列车数={source_count} 实验未完成（退出状态{return_code}）"
                )
        validate_source_count(case_dir, source_count)
        case_rows = collect_case_result(args.seed, source_count, case_dir)
        all_rows.extend(case_rows)
        write_outputs(root_dir, all_rows, source_counts)

    print(f"完成 {len(source_counts)} 组{EXPERIMENT_NAME}")


if __name__ == "__main__":
    main()
