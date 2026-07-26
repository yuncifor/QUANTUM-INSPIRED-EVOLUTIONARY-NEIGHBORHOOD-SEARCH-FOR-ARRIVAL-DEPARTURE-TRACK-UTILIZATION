from __future__ import annotations

import argparse
import csv
import os
import shutil
import statistics
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None
    Border = Font = PatternFill = Side = None
    get_column_letter = None
    EXCEL_AVAILABLE = False


BASE_SCRIPT_NAME = "run_frankfurt_hbf_benchmark.py"
DEFAULT_OUTPUT_DIR = "frankfurt_hbf_qea_ns_vs_cp_sat_multi_seed"
METHOD_ORDER = ["QEA-NS", "CP-SAT"]
SUMMARY_FILE = "法兰克福中央车站QEA-NS与CP-SAT对比实验汇总.csv"
DETAIL_FILE = "法兰克福中央车站QEA-NS与CP-SAT方案指标对比.csv"
PERFORMANCE_FILE = "法兰克福中央车站QEA-NS与CP-SAT求解性能对比.csv"
REQUIRED_OUTPUT_FILES = (SUMMARY_FILE, DETAIL_FILE, PERFORMANCE_FILE)
DETAIL_FIELDS = [
    "seed",
    "方法",
    "状态",
    "总冲突项",
    "硬冲突项",
    "进站咽喉冲突项",
    "股道占用冲突项",
    "出站咽喉冲突项",
    "总晚点时长",
    "平均晚点时长",
    "最大晚点时长",
    "能量值",
    "求解时间(s)",
    "QEA-NS候选解数量",
    "内生顺延列车数",
    "内生顺延合计(min)",
    "内生最大顺延(min)",
]


def parse_seed_spec(seed_spec: str) -> list[int]:
    seeds: list[int] = []
    for part in seed_spec.split(","):
        token = part.strip()
        if not token:
            continue
        if "-" in token:
            start_text, end_text = token.split("-", 1)
            start = int(start_text.strip())
            end = int(end_text.strip())
            step = 1 if end >= start else -1
            seeds.extend(range(start, end + step, step))
        else:
            seeds.append(int(token))
    return list(dict.fromkeys(seeds))


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def to_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "--":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def read_detail_metrics(seed_dir: Path) -> dict[str, dict[str, str]]:
    detail_rows = read_csv_rows(seed_dir / DETAIL_FILE)
    metrics_by_method = {method: {} for method in METHOD_ORDER}
    for row in detail_rows:
        metric = row.get("指标", "")
        for method in METHOD_ORDER:
            metrics_by_method[method][metric] = row.get(method, "--")
    return metrics_by_method


def read_performance_metrics(seed_dir: Path) -> dict[str, dict[str, str]]:
    performance_rows = read_csv_rows(seed_dir / PERFORMANCE_FILE)
    return {row.get("方法", ""): row for row in performance_rows}


def collect_seed_result(seed: int, seed_dir: Path) -> list[dict[str, object]]:
    detail = read_detail_metrics(seed_dir)
    performance = read_performance_metrics(seed_dir)
    rows: list[dict[str, object]] = []
    for method in METHOD_ORDER:
        method_detail = detail.get(method, {})
        method_perf = performance.get(method, {})
        rows.append(
            {
                "seed": seed,
                "方法": method,
                "状态": method_perf.get("状态", "--"),
                "总冲突项": method_detail.get("总冲突项", "--"),
                "硬冲突项": method_detail.get("硬冲突项", "--"),
                "进站咽喉冲突项": method_detail.get("进站咽喉冲突项", "--"),
                "股道占用冲突项": method_detail.get("股道占用冲突项", "--"),
                "出站咽喉冲突项": method_detail.get("出站咽喉冲突项", "--"),
                "总晚点时长": method_detail.get("总晚点时长", "--"),
                "平均晚点时长": method_detail.get("平均晚点时长", "--"),
                "最大晚点时长": method_detail.get("最大晚点时长", "--"),
                "能量值": method_detail.get("能量值", "--"),
                "求解时间(s)": method_perf.get("求解时间(s)", "--"),
                "QEA-NS候选解数量": method_perf.get("QEA-NS候选解数量", "--"),
                "内生顺延列车数": method_perf.get("内生顺延列车数", "--"),
                "内生顺延合计(min)": method_perf.get("内生顺延合计(min)", "--"),
                "内生最大顺延(min)": method_perf.get("内生最大顺延(min)", "--"),
            }
        )
    return rows


def aggregate_method(rows: list[dict[str, object]], method: str) -> dict[str, object]:
    method_rows = [row for row in rows if row.get("方法") == method]
    delays = [
        value
        for value in (to_number(row.get("总晚点时长")) for row in method_rows)
        if value is not None
    ]
    hard_conflicts = [
        value
        for value in (to_number(row.get("硬冲突项")) for row in method_rows)
        if value is not None
    ]
    times = [
        value
        for value in (to_number(row.get("求解时间(s)")) for row in method_rows)
        if value is not None
    ]
    return {
        "方法": method,
        "样本数": len(method_rows),
        "有效样本数": len(delays),
        "零硬冲突次数": sum(1 for value in hard_conflicts if value == 0),
        "总晚点平均值": round(statistics.mean(delays), 3) if delays else "--",
        "总晚点中位数": round(statistics.median(delays), 3) if delays else "--",
        "总晚点最小值": min(delays) if delays else "--",
        "总晚点最大值": max(delays) if delays else "--",
        "总晚点标准差": round(statistics.pstdev(delays), 3)
        if len(delays) > 1
        else 0
        if delays
        else "--",
        "求解时间平均值(s)": round(statistics.mean(times), 3) if times else "--",
    }


def value_for(
    rows: list[dict[str, object]], seed: int, method: str, field: str
) -> float | None:
    for row in rows:
        if row.get("seed") == seed and row.get("方法") == method:
            return to_number(row.get(field))
    return None


def build_win_rows(
    rows: list[dict[str, object]], seeds: list[int]
) -> list[dict[str, object]]:
    comparisons = [
        ("QEA-NS", "CP-SAT"),
    ]
    result_rows = []
    for left, right in comparisons:
        wins = losses = ties = comparable = 0
        margins = []
        for seed in seeds:
            left_value = value_for(rows, seed, left, "总晚点时长")
            right_value = value_for(rows, seed, right, "总晚点时长")
            if left_value is None or right_value is None:
                continue
            comparable += 1
            margin = right_value - left_value
            margins.append(margin)
            if margin > 0:
                wins += 1
            elif margin < 0:
                losses += 1
            else:
                ties += 1
        result_rows.append(
            {
                "比较": f"{left} 优于 {right}",
                "可比较seed数": comparable,
                "胜出次数": wins,
                "打平次数": ties,
                "落后次数": losses,
                "胜率": round(wins / comparable, 3) if comparable else "--",
                "平均优势(分钟)": round(statistics.mean(margins), 3)
                if margins
                else "--",
                "中位优势(分钟)": round(statistics.median(margins), 3)
                if margins
                else "--",
            }
        )
    return result_rows


def write_excel(path: Path, sheets: list[tuple[str, list[dict[str, object]]]]) -> None:
    if not EXCEL_AVAILABLE:
        return
    assert openpyxl is not None
    assert Border is not None
    assert Font is not None
    assert PatternFill is not None
    assert Side is not None
    assert get_column_letter is not None
    workbook = openpyxl.Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for index, (sheet_name, rows) in enumerate(sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        if worksheet is None:
            worksheet = workbook.create_sheet(sheet_name)
        worksheet.title = sheet_name
        if not rows:
            continue
        fieldnames = list(rows[0])
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([row.get(field, "") for field in fieldnames])
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_index = int(column_cells[0].column or 1)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12), 36
            )
    workbook.save(path)


def build_command(args: argparse.Namespace, seed: int, seed_dir: Path) -> list[str]:
    command = [
        sys.executable,
        str(Path(__file__).resolve().parent / BASE_SCRIPT_NAME),
        "--seed",
        str(seed),
        "--output-dir",
        str(seed_dir),
        "--max-route-candidates",
        str(args.max_route_candidates),
        "--qea-time-limit",
        str(args.qea_time_limit),
        "--subproblem-time-limit",
        str(args.subproblem_time_limit),
        "--cp-sat-time-limit",
        str(args.cp_sat_time_limit),
        "--qea-pop-size",
        str(args.qea_pop_size),
        "--qea-max-generations",
        str(args.qea_max_generations),
        "--qea-neighborhood-restarts",
        str(args.qea_neighborhood_restarts),
        "--safety-wait-rounds",
        str(args.safety_wait_rounds),
    ]
    if args.short_case_train_limit:
        command.extend(["--short-case-train-limit", str(args.short_case_train_limit)])
    if args.input:
        command.extend(["--input", args.input])
    return command


def seed_completed(seed_dir: Path) -> bool:
    if not all((seed_dir / filename).exists() for filename in REQUIRED_OUTPUT_FILES):
        return False
    detail_rows = read_csv_rows(seed_dir / DETAIL_FILE)
    performance_rows = read_csv_rows(seed_dir / PERFORMANCE_FILE)
    summary_rows = read_csv_rows(seed_dir / SUMMARY_FILE)
    detail_fields = set(detail_rows[0]) if detail_rows else set()
    performance_methods = {row.get("方法", "") for row in performance_rows}
    summary_items = {row.get("项目", "") for row in summary_rows}
    return (
        set(METHOD_ORDER).issubset(detail_fields)
        and set(METHOD_ORDER).issubset(performance_methods)
        and {"QEA-NS总晚点", "CP-SAT总晚点"}.issubset(summary_items)
    )


def cleanup_seed_outputs(seed_dir: Path) -> None:
    for path in seed_dir.iterdir():
        if path.is_dir():
            shutil.rmtree(path)
        else:
            path.unlink()


def run_seed_command(command: list[str]) -> int:
    env = os.environ.copy()
    env["PYTHONHASHSEED"] = "0"
    process = subprocess.Popen(
        command,
        cwd=Path(__file__).resolve().parent,
        stderr=subprocess.STDOUT,
        text=True,
        env=env,
    )
    try:
        return process.wait()
    except KeyboardInterrupt:
        print("用户中断，正在终止当前随机种子实验。")
        process.terminate()
        try:
            process.wait(timeout=30)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait()
        return 130


def write_aggregate_outputs(
    root_dir: Path, rows: list[dict[str, object]], seeds: list[int]
) -> None:
    summary_rows = [aggregate_method(rows, method) for method in METHOD_ORDER]
    win_rows = build_win_rows(rows, seeds)
    write_csv(root_dir / "多随机种子逐次结果.csv", rows, DETAIL_FIELDS)
    write_csv(root_dir / "多随机种子统计汇总.csv", summary_rows, list(summary_rows[0]))
    write_csv(root_dir / "多随机种子胜率统计.csv", win_rows, list(win_rows[0]))
    write_excel(
        root_dir / "多随机种子实验汇总.xlsx",
        [
            ("逐seed结果", rows),
            ("统计汇总", summary_rows),
            ("胜率统计", win_rows),
        ],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="法兰克福中央车站多随机种子实验")
    parser.add_argument("--seeds", type=str, default="42-51")
    parser.add_argument("--input", type=str, default="")
    parser.add_argument("--output-dir", type=str, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--keep-existing", action="store_true")
    parser.add_argument("--rerun-completed", action="store_true")
    parser.add_argument("--max-route-candidates", type=int, default=5)
    parser.add_argument("--qea-time-limit", type=float, default=400.0)
    parser.add_argument("--subproblem-time-limit", type=float, default=120.0)
    parser.add_argument(
        "--cp-sat-time-limit",
        dest="cp_sat_time_limit",
        type=float,
        default=120.0,
    )
    parser.add_argument("--qea-pop-size", type=int, default=50)
    parser.add_argument("--qea-max-generations", type=int, default=500)
    parser.add_argument("--qea-neighborhood-restarts", type=int, default=1)
    parser.add_argument("--safety-wait-rounds", type=int, default=160)
    parser.add_argument("--short-case-train-limit", type=int, default=0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    seeds = parse_seed_spec(args.seeds)
    if not seeds:
        raise ValueError("至少需要一个随机种子。")

    root_dir = Path(args.output_dir).expanduser().resolve()
    root_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, object]] = []

    for seed in seeds:
        seed_dir = root_dir / f"seed_{seed}"
        seed_dir.mkdir(parents=True, exist_ok=True)
        completed_code = 0
        if seed_completed(seed_dir) and not args.rerun_completed:
            print(f"seed={seed} 已完成，跳过。")
        else:
            if not args.keep_existing:
                cleanup_seed_outputs(seed_dir)
            command = build_command(args, seed, seed_dir)
            print(f"开始 seed={seed}")
            completed_code = run_seed_command(command)
            if completed_code == 130:
                print(f"seed={seed} 被中断，已保留此前完成结果。")
                seed_rows = collect_seed_result(seed, seed_dir)
                all_rows.extend(seed_rows)
                write_aggregate_outputs(root_dir, all_rows, seeds)
                break
        seed_rows = collect_seed_result(seed, seed_dir)
        all_rows.extend(seed_rows)
        write_aggregate_outputs(root_dir, all_rows, seeds)

    print(f"完成 {len(seeds)} 个随机种子实验")


if __name__ == "__main__":
    main()
